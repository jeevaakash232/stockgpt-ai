"""
History API
-----------
GET /api/history/{symbol}              — Daily history for a symbol (last 30 days)
GET /api/history/{symbol}/intraday     — Today's intraday ticks
GET /api/history/dates                 — All available trading dates
GET /api/history/snapshot/{date}       — All symbols for a specific date
GET /api/history/stats                 — Database statistics
GET /api/history/download/{date}       — Download Excel for a specific date
"""

import io
from datetime import datetime
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse

from app.services.history_service import (
    get_daily_history,
    get_intraday_ticks,
    get_all_symbols_history,
    get_available_dates,
    get_db_stats,
)

router = APIRouter()


@router.get("/history/stats", summary="Database statistics")
def history_stats():
    return get_db_stats()


@router.get("/history/dates", summary="Available trading dates")
def history_dates():
    return get_available_dates()


@router.get("/history/{symbol}", summary="Daily history for a symbol")
def symbol_history(
    symbol: str,
    days: int = Query(default=30, ge=1, le=365),
):
    """Returns daily snapshots for the given symbol over the last N days."""
    try:
        data = get_daily_history(symbol.upper(), days=days)
        return {"symbol": symbol.upper(), "days": days, "records": data}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/history/{symbol}/intraday", summary="Today's intraday ticks")
def symbol_intraday(
    symbol: str,
    date: str = Query(default=None, description="YYYY-MM-DD, defaults to today"),
):
    try:
        data = get_intraday_ticks(symbol.upper(), trade_date=date)
        return {"symbol": symbol.upper(), "date": date, "ticks": data}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/history/snapshot/{date}", summary="All symbols for a date")
def date_snapshot(date: str):
    """Returns all symbols' data for a specific trading date (YYYY-MM-DD)."""
    try:
        data = get_all_symbols_history(trade_date=date)
        if not data:
            raise HTTPException(
                status_code=404,
                detail=f"No data found for {date}. Available dates: /api/history/dates"
            )
        return {"date": date, "count": len(data), "records": data}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/history/download/{date}", summary="Download Excel for a date")
def download_history_excel(date: str):
    """
    Download a formatted Excel file with all stock data for a specific trading date.
    Date format: YYYY-MM-DD  (e.g. 2026-07-05)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        rows = get_all_symbols_history(trade_date=date)
        if not rows:
            raise HTTPException(status_code=404, detail=f"No data for {date}")

        wb  = Workbook()
        wb.remove(wb.active)
        ws  = wb.create_sheet(f"Snapshot {date}")
        ws.freeze_panes = "A3"

        HDR_FILL = PatternFill("solid", fgColor="1F4E79")
        GRN_FILL = PatternFill("solid", fgColor="E2EFDA")
        RED_FILL = PatternFill("solid", fgColor="FCE4D6")
        WHT_FILL = PatternFill("solid", fgColor="FFFFFF")
        HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        BODY     = Font(name="Calibri", size=11)
        INFO     = Font(name="Calibri", italic=True, color="595959", size=10)
        _side    = Side(style="thin", color="BFBFBF")
        BORDER   = Border(left=_side, right=_side, top=_side, bottom=_side)

        def _h(ws, row, cols):
            for c, h in enumerate(cols, 1):
                cell = ws.cell(row=row, column=c, value=h)
                cell.font = HDR_FONT; cell.fill = HDR_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = BORDER

        def _r(ws, row_n, values, fill=None):
            for c, v in enumerate(values, 1):
                cell = ws.cell(row=row_n, column=c, value=v)
                cell.font = BODY; cell.fill = fill or WHT_FILL
                cell.border = BORDER
                cell.alignment = Alignment(
                    horizontal="right" if isinstance(v, (int, float)) else "left",
                    vertical="center"
                )

        # Info row
        info_cell = ws.cell(row=1, column=1,
            value=f"StockGPT AI — Historical Snapshot   |   Date: {date}   |   {len(rows)} symbols")
        info_cell.font = INFO

        # Headers
        headers = ["#", "Symbol", "LTP ₹", "Prev Close ₹", "Price Δ%",
                   "PCR", "Signal", "Call OI", "Put OI", "Max Pain ₹"]
        _h(ws, 2, headers)

        for i, s in enumerate(rows):
            r    = i + 3
            sig  = (s.get("signal") or "").lower()
            fill = GRN_FILL if "bullish" in sig else (RED_FILL if "bearish" in sig else WHT_FILL)
            pct  = s.get("price_chg_pct")

            _r(ws, r, [
                i + 1,
                s["symbol"],
                s.get("ltp"),
                s.get("prev_close"),
                pct,
                s.get("pcr"),
                s.get("signal"),
                s.get("call_oi"),
                s.get("put_oi"),
                s.get("max_pain"),
            ], fill)

            for col, fmt in [(3, "#,##0.00"), (4, "#,##0.00"), (5, "+0.0;-0.0"),
                             (6, "0.00"), (8, "#,##0"), (9, "#,##0"), (10, "#,##0.00")]:
                ws.cell(row=r, column=col).number_format = fmt

        # Column widths
        for i, w in enumerate([4, 16, 13, 14, 11, 7, 16, 14, 14, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"StockGPT_History_{date}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Cache-Control": "no-cache, no-store, must-revalidate",
                "Pragma": "no-cache",
                "Expires": "0"
            },
        )

    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
