"""
Export API
----------
GET /api/export/excel  — Download all NSE F&O stocks as clean editable Excel
GET /api/export/csv    — Download all stocks as plain CSV

Excel format:
  - White background, Calibri font, standard cell sizes
  - Blue header row, light green/red row tints (fully editable)
  - No merged cells (except info row), no locked cells
  - 6 sheets: All Stocks, Top Gainers, Top Losers, Most Active, Indices, Watchlist
"""

import io
from datetime import datetime
from fastapi import APIRouter
from fastapi.responses import StreamingResponse

router = APIRouter()


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

@router.get("/export/excel", summary="Download all stock data as Excel")
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Data ──────────────────────────────────────────────────
    from app.services.market_data    import _build_market
    from app.services.market_service import (
        get_top_gainers, get_top_losers, get_most_active, get_indices
    )
    from app.services import watchlist_service

    # Bypass cache to fetch absolute latest values from API/DB
    from app.services import cache_service
    cache_service.invalidate("prev_day_snapshot")
    cache_service.invalidate("market_data")
    market  = _build_market()
    gainers = get_top_gainers()
    losers  = get_top_losers()
    active  = get_most_active()
    watchl  = watchlist_service.get_watchlist_with_prices()
    indices = get_indices()

    now   = datetime.now()
    stamp = now.strftime("%d %b %Y  %H:%M:%S")

    # ── Shared styles ─────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark blue
    GRN_FILL  = PatternFill("solid", fgColor="E2EFDA")   # light green
    RED_FILL  = PatternFill("solid", fgColor="FCE4D6")   # light red/orange
    WHT_FILL  = PatternFill("solid", fgColor="FFFFFF")

    HDR_FONT  = Font(name="Calibri", bold=True,  color="FFFFFF", size=11)
    BODY_FONT = Font(name="Calibri", bold=False, color="000000", size=11)
    INFO_FONT = Font(name="Calibri", italic=True, color="595959", size=10)
    BOLD_GRN  = Font(name="Calibri", bold=True,  color="375623", size=11)
    BOLD_RED  = Font(name="Calibri", bold=True,  color="9C0006", size=11)
    BOLD_BLU  = Font(name="Calibri", bold=True,  color="203864", size=11)

    _side    = Side(style="thin", color="BFBFBF")
    BORDER   = Border(left=_side, right=_side, top=_side, bottom=_side)
    CENTER   = Alignment(horizontal="center", vertical="center")
    LEFT     = Alignment(horizontal="left",   vertical="center")
    RIGHT    = Alignment(horizontal="right",  vertical="center")

    def _widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _header(ws, row, cols):
        for c, h in enumerate(cols, 1):
            cell            = ws.cell(row=row, column=c, value=h)
            cell.font       = HDR_FONT
            cell.fill       = HDR_FILL
            cell.alignment  = CENTER
            cell.border     = BORDER
        ws.row_dimensions[row].height = 18

    def _info(ws, row, text):
        cell           = ws.cell(row=row, column=1, value=text)
        cell.font      = INFO_FONT
        cell.alignment = LEFT
        ws.row_dimensions[row].height = 15

    def _row(ws, row_n, values, fill=None):
        for c, v in enumerate(values, 1):
            cell           = ws.cell(row=row_n, column=c, value=v)
            cell.font      = BODY_FONT
            cell.fill      = fill or WHT_FILL
            cell.border    = BORDER
            cell.alignment = RIGHT if isinstance(v, (int, float)) else LEFT

    def _signal_fill(sig_str):
        s = (sig_str or "").lower()
        if "bullish" in s: return GRN_FILL
        if "bearish" in s: return RED_FILL
        return WHT_FILL

    # ── Workbook ──────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    # ════════════════════════════════════════════════════════
    # Sheet 1 — All Stocks (209 NSE F&O)
    # ════════════════════════════════════════════════════════
    ws = wb.create_sheet("All Stocks")
    ws.freeze_panes = "A3"

    _info(ws, 1, f"StockGPT AI — NSE F&O Live Market Data   |   Generated: {stamp}   |   {len(market)} stocks")
    _header(ws, 2, ["#", "Symbol", "Price (₹)", "Price Change %", "Call OI", "Put OI", "Max Pain (₹)", "Current Day PCR", "Previous Day PCR", "Δ PCR", "% Change in PCR", "Expiry Date", "Signal"])

    for i, s in enumerate(market):
        r    = i + 3
        fill = _signal_fill(s["signal"])

        price_chg = s.get("price_chg_pct")
        price_chg_val = price_chg / 100.0 if price_chg is not None else None

        pcr_chg_pct = s.get("pcr_change_pct")

        _row(ws, r, [
            i+1, 
            s["symbol"], 
            s["ltp"], 
            price_chg_val,
            s["call_oi"], 
            s["put_oi"], 
            s["max_pain"],
            s["pcr"],
            s.get("prev_day_pcr"),
            s.get("pcr_change"),
            pcr_chg_pct,
            s.get("expiry"),
            s["signal"]
        ], fill)

        # Bold coloured text for signal column (column 13)
        sig_cell = ws.cell(row=r, column=13)
        sig_lower = s["signal"].lower()
        if "bullish" in sig_lower:   sig_cell.font = BOLD_GRN
        elif "bearish" in sig_lower: sig_cell.font = BOLD_RED
        else:                        sig_cell.font = BOLD_BLU

        ws.cell(row=r, column=3).number_format = '#,##0.00'
        ws.cell(row=r, column=4).number_format = '+0.0%;-0.0%;0.0%'
        ws.cell(row=r, column=5).number_format = '#,##0'
        ws.cell(row=r, column=6).number_format = '#,##0'
        ws.cell(row=r, column=7).number_format = '#,##0.00'
        ws.cell(row=r, column=8).number_format = '0.00'
        ws.cell(row=r, column=9).number_format = '0.00'
        ws.cell(row=r, column=10).number_format = '+0.00;-0.00;0.00'
        ws.cell(row=r, column=11).number_format = '+0.00;-0.00;0.00'

    _widths(ws, [4, 16, 14, 15, 14, 14, 15, 16, 17, 10, 17, 14, 16])

    # ════════════════════════════════════════════════════════
    # Sheet 2 — Top Gainers
    # ════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Top Gainers")
    ws2.freeze_panes = "A3"
    _info(ws2, 1, f"Top Gainers by % Change   |   {stamp}")
    _header(ws2, 2, ["#", "Symbol", "Price (₹)", "Change %", "High (₹)", "Low (₹)", "Volume"])
    for i, s in enumerate(gainers):
        r = i + 3
        _row(ws2, r, [i+1, s["symbol"], s["price"], s["change_pct"],
                      s.get("high"), s.get("low"), s["volume"]], GRN_FILL)
        ws2.cell(row=r, column=3).number_format = '#,##0.00'
        ws2.cell(row=r, column=4).number_format = '+0.00;-0.00'
        ws2.cell(row=r, column=5).number_format = '#,##0.00'
        ws2.cell(row=r, column=6).number_format = '#,##0.00'
        ws2.cell(row=r, column=7).number_format = '#,##0'
    _widths(ws2, [4, 16, 13, 11, 13, 13, 14])

    # ════════════════════════════════════════════════════════
    # Sheet 3 — Top Losers
    # ════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Top Losers")
    ws3.freeze_panes = "A3"
    _info(ws3, 1, f"Top Losers by % Change   |   {stamp}")
    _header(ws3, 2, ["#", "Symbol", "Price (₹)", "Change %", "High (₹)", "Low (₹)", "Volume"])
    for i, s in enumerate(losers):
        r = i + 3
        _row(ws3, r, [i+1, s["symbol"], s["price"], s["change_pct"],
                      s.get("high"), s.get("low"), s["volume"]], RED_FILL)
        ws3.cell(row=r, column=3).number_format = '#,##0.00'
        ws3.cell(row=r, column=4).number_format = '+0.00;-0.00'
        ws3.cell(row=r, column=5).number_format = '#,##0.00'
        ws3.cell(row=r, column=6).number_format = '#,##0.00'
        ws3.cell(row=r, column=7).number_format = '#,##0'
    _widths(ws3, [4, 16, 13, 11, 13, 13, 14])

    # ════════════════════════════════════════════════════════
    # Sheet 4 — Most Active
    # ════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Most Active")
    ws4.freeze_panes = "A3"
    _info(ws4, 1, f"Most Active by Volume   |   {stamp}")
    _header(ws4, 2, ["#", "Symbol", "Price (₹)", "Change %", "Volume"])
    for i, s in enumerate(active):
        r    = i + 3
        pct  = s.get("change_pct", 0) or 0
        fill = GRN_FILL if pct >= 0 else RED_FILL
        _row(ws4, r, [i+1, s["symbol"], s["price"], pct, s["volume"]], fill)
        ws4.cell(row=r, column=3).number_format = '#,##0.00'
        ws4.cell(row=r, column=4).number_format = '+0.00;-0.00'
        ws4.cell(row=r, column=5).number_format = '#,##0'
    _widths(ws4, [4, 16, 13, 11, 16])

    # ════════════════════════════════════════════════════════
    # Sheet 5 — Indices
    # ════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Indices")
    ws5.freeze_panes = "A3"
    _info(ws5, 1, f"Live Indices   |   {stamp}")
    _header(ws5, 2, ["Index", "Price (₹)", "Change (₹)", "Change %", "High", "Low"])
    for i, (name, d) in enumerate(indices.items()):
        r    = i + 3
        pct  = d.get("change_pct", 0) or 0
        fill = GRN_FILL if pct >= 0 else RED_FILL
        _row(ws5, r, [name, d.get("current_price"), d.get("change"), pct,
                      d.get("high"), d.get("low")], fill)
        for col in [2, 3, 5, 6]:
            ws5.cell(row=r, column=col).number_format = '#,##0.00'
        ws5.cell(row=r, column=4).number_format = '+0.00;-0.00'
    _widths(ws5, [14, 14, 13, 12, 12, 12])

    # ════════════════════════════════════════════════════════
    # Sheet 6 — Watchlist
    # ════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Watchlist")
    ws6.freeze_panes = "A3"
    _info(ws6, 1, f"My Watchlist   |   {stamp}")
    _header(ws6, 2, ["Symbol", "Price (₹)", "Change %"])
    if watchl:
        for i, s in enumerate(watchl):
            r    = i + 3
            pct  = s.get("change_pct", 0) or 0
            fill = GRN_FILL if pct >= 0 else RED_FILL
            _row(ws6, r, [s["symbol"], s.get("price"), pct], fill)
            ws6.cell(row=r, column=2).number_format = '#,##0.00'
            ws6.cell(row=r, column=3).number_format = '+0.00;-0.00'
    else:
        ws6.cell(row=3, column=1, value="Watchlist is empty").font = INFO_FONT
    _widths(ws6, [18, 15, 13])

    # ════════════════════════════════════════════════════════
    # Sheet 7 — Yesterday's Historical Snapshot (if available)
    # ════════════════════════════════════════════════════════
    try:
        from app.services.history_service import get_all_symbols_history, get_available_dates
        from datetime import date as _date, timedelta

        dates     = get_available_dates()
        hist_date = None
        hist_rows = []

        # Try yesterday first, then most recent available date
        yesterday = (_date.today() - timedelta(days=1)).isoformat()
        if yesterday in dates:
            hist_date = yesterday
        elif dates:
            hist_date = dates[0]   # most recent

        if hist_date:
            hist_rows = get_all_symbols_history(trade_date=hist_date)

        if hist_rows:
            ws7 = wb.create_sheet(f"History {hist_date}")
            ws7.freeze_panes = "A3"
            _info(ws7, 1, f"Historical Snapshot   |   {hist_date}   |   {len(hist_rows)} symbols")
            _header(ws7, 2, ["#", "Symbol", "LTP (₹)", "Prev Close (₹)", "Price Δ%",
                              "PCR", "Signal", "Call OI", "Put OI", "Max Pain (₹)"])

            for i, s in enumerate(hist_rows):
                r    = i + 3
                fill = _signal_fill(s.get("signal", ""))
                pct  = s.get("price_chg_pct")
                _row(ws7, r, [
                    i + 1,
                    s["symbol"],
                    s.get("ltp"),
                    s.get("prev_close"),
                    pct,
                    s.get("pcr"),
                    s.get("signal"),
                    s.get("call_oi"),
                    s.get("put_oi"),
                    s.get("max_pain"),
                ], fill)
                for col, fmt in [(3, "#,##0.00"), (4, "#,##0.00"), (5, "+0.0;-0.0"),
                                 (6, "0.00"), (8, "#,##0"), (9, "#,##0"), (10, "#,##0.00")]:
                    ws7.cell(row=r, column=col).number_format = fmt

            _widths(ws7, [4, 16, 13, 14, 11, 7, 16, 14, 14, 14])
    except Exception:
        pass   # history not yet populated — skip this sheet silently

    # ── Stream ────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"StockGPT_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0"
        },
    )


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------

@router.get("/export/csv", summary="Download all stocks as CSV")
def export_csv():
    """Plain CSV of all NSE F&O stocks — opens directly in Excel or Google Sheets."""
    from app.services.market_data import _build_market

    # Bypass cache to fetch absolute latest values
    from app.services import cache_service
    cache_service.invalidate("prev_day_snapshot")
    cache_service.invalidate("market_data")
    market = _build_market()
    lines  = ["Symbol,Price,Price Change %,Call OI,Put OI,Max Pain,Current Day PCR,Previous Day PCR,Delta PCR,PCR Change %,Expiry Date,Signal"]
    for s in market:
        price_chg = f"{s['price_chg_pct']}%" if s.get('price_chg_pct') is not None else ""
        pcr_chg_pct = s.get('pcr_change_pct') if s.get('pcr_change_pct') is not None else ""
        prev_pcr = s.get('prev_day_pcr') if s.get('prev_day_pcr') is not None else ""
        pcr_change = s.get('pcr_change') if s.get('pcr_change') is not None else ""
        expiry = s.get('expiry') if s.get('expiry') is not None else ""

        lines.append(
            f"{s['symbol']},{s['ltp']},{price_chg},"
            f"{s['call_oi']},{s['put_oi']},{s['max_pain']},"
            f"{s['pcr']},{prev_pcr},{pcr_change},{pcr_chg_pct},{expiry},{s['signal']}"
        )

    filename = f"StockGPT_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return StreamingResponse(
        io.StringIO("\n".join(lines)),
        media_type="text/csv",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0"
        },
    )
